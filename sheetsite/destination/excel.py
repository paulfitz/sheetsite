def write_destination_excel(params, state):
    workbook = state['workbook']
    output_file = params['output_file']
    from openpyxl import Workbook
    wb = Workbook()
    first = True
    for sheet in workbook.worksheets():
        title = sheet.title
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = title
        rows = sheet.get_all_values()
        for r, row in enumerate(rows):
            for c, cell in enumerate(row):
                ws.cell(row=r+1, column=c+1).value = cell
    wb.save(output_file)
    return True
